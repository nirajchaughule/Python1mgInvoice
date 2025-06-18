import imaplib
import email
import os
import re
import time
import traceback
import pdfplumber
from email.header import decode_header
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import hashlib
import html2text  # For converting HTML to plain text

# Configuration
EMAIL = "youremail"
PASSWORD = "yourapppassword"
IMAP_SERVER = "imap.gmail.com"
SAVE_DIR = "1mg_Invoices"
os.makedirs(SAVE_DIR, exist_ok=True)

# Track processed orders and files to prevent duplicates
processed_orders = set()
processed_files = {}

def clean_filename(text):
    """Sanitize filenames to avoid invalid characters."""
    if text is None:
        return "unknown"
    return "".join(c if c.isalnum() or c in (' ', '.', '_') else "_" for c in text)

def extract_subtotal(text):
    """Enhanced Subtotal extraction with flexible patterns."""
    if not text:
        return None
        
    # More flexible patterns to match different formats
    patterns = [
        r"Subtotal\b[\s:]*Rs?\.?\s*(\d{1,3}(?:,\d{3})*\.\d{2})",  # "Subtotal Rs.48.50"
        r"Subtotal\b[\s:]*INR\s*(\d{1,3}(?:,\d{3})*\.\d{2})",     # "Subtotal INR 48.50"
        r"Subtotal\b[\s:]*₹?\s*(\d{1,3}(?:,\d{3})*\.\d{2})",       # "Subtotal ₹48.50"
        r"Total\s+Items\b[\s:]*Rs?\.?\s*(\d{1,3}(?:,\d{3})*\.\d{2})",  # "Total Items Rs.48.50"
        r"Item\s+Total\b[\s:]*Rs?\.?\s*(\d{1,3}(?:,\d{3})*\.\d{2})",   # "Item Total Rs.48.50"
        r"Total\s+Products\b[\s:]*Rs?\.?\s*(\d{1,3}(?:,\d{3})*\.\d{2})",  # "Total Products Rs.48.50"
        r"Subtotal\b[^\d]*(\d+\.\d{2})",  # Fallback pattern
        r"Rs?\.?\s*(\d+\.\d{2})[\s\S]*Subtotal"  # Reverse pattern
    ]
    
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            amount_str = match.group(1).replace(',', '')
            try:
                return float(amount_str)
            except ValueError:
                continue
    return None

def html_to_plain_text(html_content):
    """Convert HTML email content to clean plain text."""
    if not html_content:
        return ""
    
    # Create HTML to text converter
    h = html2text.HTML2Text()
    h.ignore_links = True
    h.ignore_images = True
    h.ignore_tables = True
    h.ignore_emphasis = True
    h.body_width = 0  # Don't wrap text
    
    # Convert HTML to plain text
    plain_text = h.handle(html_content)
    
    # Remove excessive blank lines
    plain_text = re.sub(r'\n\s*\n', '\n\n', plain_text)
    
    return plain_text.strip()

def create_pdf_from_body(order_id, subject, body, filename):
    """Create a PDF from email body content with proper text formatting."""
    try:
        c = canvas.Canvas(filename, pagesize=letter)
        width, height = letter
        
        # Add header
        c.setFont("Helvetica-Bold", 16)
        c.drawString(72, height - 72, "1MG ORDER INVOICE")
        c.line(72, height - 80, width - 72, height - 80)
        
        # Add order details
        c.setFont("Helvetica", 12)
        y_position = height - 100
        c.drawString(72, y_position, f"Order ID: {order_id}")
        y_position -= 20
        c.drawString(72, y_position, f"Subject: {subject[:100]}{'...' if len(subject) > 100 else ''}")
        y_position -= 20
        c.drawString(72, y_position, f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        y_position -= 40
        
        # Add body content header
        c.drawString(72, y_position, "Original Email Content:")
        y_position -= 20
        c.setFont("Helvetica", 10)
        
        # Split body into lines and process each line
        lines = body.split('\n')
        for line in lines:
            # Skip empty lines
            if not line.strip():
                continue
                
            # Wrap long lines
            while line:
                # Check if we need a new page
                if y_position < 100:
                    c.showPage()
                    y_position = height - 50
                    c.setFont("Helvetica", 10)
                
                # Take up to 100 characters
                chunk = line[:100]
                line = line[100:]
                
                # Draw the text
                c.drawString(72, y_position, chunk)
                y_position -= 12
        
        # Save PDF
        c.save()
        return True
    except Exception as e:
        print(f"   ⚠️ PDF creation error: {str(e)}")
        traceback.print_exc()
        return False

def get_order_id(subject, body):
    """Extract order ID from subject or body."""
    patterns = [
        r"Order\s*#?\s*([A-Z0-9-]{6,})",
        r"Order\s+ID\s*[:]?\s*([A-Z0-9-]{6,})",
        r"Order\s+No\.?\s*[:]?\s*([A-Z0-9-]{6,})",
        r"Your\s+Order\s+([A-Z0-9-]{6,})",
        r"Order\s+([A-Z0-9-]{6,})"
    ]
    
    # First check subject
    for pattern in patterns:
        match = re.search(pattern, subject, re.IGNORECASE)
        if match:
            return match.group(1)
    
    # Then check body
    if body:
        for pattern in patterns:
            match = re.search(pattern, body, re.IGNORECASE)
            if match:
                return match.group(1)
    
    return None

def fetch_1mg_emails():
    """Fetch and process 1mg emails to create invoices."""
    print("🚀 Connecting to IMAP server...")
    try:
        mail = imaplib.IMAP4_SSL(IMAP_SERVER)
        mail.login(EMAIL, PASSWORD)
        mail.select("inbox")
    except Exception as e:
        print(f"✖ IMAP connection failed: {str(e)}")
        return []

    try:
        print("🔍 Searching for 1mg emails...")
        status, messages = mail.search(None, '(FROM "no-reply@mail.1mg.com")')
        email_ids = messages[0].split()
        total_emails = len(email_ids)
        print(f"📬 Found {total_emails} emails from 1mg")
    except Exception as e:
        print(f"✖ Search failed: {str(e)}")
        mail.logout()
        return []

    invoice_data = []
    processed_count = 0

    for i, email_id in enumerate(email_ids, 1):
        email_uid = email_id.decode() if isinstance(email_id, bytes) else str(email_id)
        print(f"\n📩 Processing email {i}/{total_emails} (UID: {email_uid})")
        start_time = time.time()
        
        try:
            # Fetch email
            status, msg_data = mail.fetch(email_id, "(RFC822)")
            if not msg_data or not isinstance(msg_data[0], tuple):
                print("   ⚠️ Empty or invalid message data")
                continue
                
            raw_email = msg_data[0][1]
            message = email.message_from_bytes(raw_email)

            # Decode subject safely
            subject = "No Subject"
            try:
                subj, encoding = decode_header(message.get("Subject", ""))[0]
                if isinstance(subj, bytes):
                    subject = subj.decode(encoding or "utf-8", errors="replace")
                else:
                    subject = str(subj)
            except Exception as e:
                print(f"   ⚠️ Subject decoding error: {str(e)}")
                
            print(f"   Subject: {subject[:80]}{'...' if len(subject) > 80 else ''}")

            # Extract email body - prefer plain text, fallback to HTML
            plain_body = ""
            html_body = ""
            
            if message.is_multipart():
                for part in message.walk():
                    content_type = part.get_content_type()
                    content_disposition = str(part.get("Content-Disposition"))
                    
                    # Skip attachments
                    if "attachment" in content_disposition:
                        continue
                        
                    if content_type == "text/plain":
                        # Found plain text part
                        payload = part.get_payload(decode=True)
                        if payload:
                            try:
                                plain_body = payload.decode('utf-8', errors='replace')
                            except UnicodeDecodeError:
                                try:
                                    plain_body = payload.decode('latin-1', errors='replace')
                                except:
                                    pass
                    
                    elif content_type == "text/html":
                        # Found HTML part
                        payload = part.get_payload(decode=True)
                        if payload:
                            try:
                                html_body = payload.decode('utf-8', errors='replace')
                            except UnicodeDecodeError:
                                try:
                                    html_body = payload.decode('latin-1', errors='replace')
                                except:
                                    pass
            else:
                # Not multipart - single part email
                payload = message.get_payload(decode=True)
                if payload:
                    content_type = message.get_content_type()
                    if content_type == "text/plain":
                        try:
                            plain_body = payload.decode('utf-8', errors='replace')
                        except UnicodeDecodeError:
                            try:
                                plain_body = payload.decode('latin-1', errors='replace')
                            except:
                                pass
                    elif content_type == "text/html":
                        try:
                            html_body = payload.decode('utf-8', errors='replace')
                        except UnicodeDecodeError:
                            try:
                                html_body = payload.decode('latin-1', errors='replace')
                            except:
                                pass
            
            # Convert HTML to plain text if we didn't get a plain text version
            if plain_body:
                body_for_processing = plain_body
                body_for_pdf = plain_body
            elif html_body:
                body_for_processing = html_to_plain_text(html_body)
                body_for_pdf = body_for_processing
            else:
                body_for_processing = ""
                body_for_pdf = ""
            
            # Clean up body text for processing
            cleaned_body = re.sub(r'\s+', ' ', body_for_processing).strip()
            
            # Extract order ID
            order_id = get_order_id(subject, cleaned_body)
            if not order_id:
                print("   ⚠️ Order ID not found - skipping")
                continue
            
            # Ensure unique order ID for created PDFs
            if order_id in processed_orders:
                print(f"   ⚠️ Duplicate order {order_id}, skipping")
                continue
                
            print(f"   Order ID: {order_id}")

            # Process PDF attachments
            pdf_processed = False
            amount = None
            pdf_path = None
            
            for part in message.walk():
                if part.get_content_type() == "application/pdf":
                    filename = part.get_filename()
                    if filename:
                        try:
                            payload = part.get_payload(decode=True)
                            if not payload:
                                print(f"   ⚠️ Empty PDF payload: {filename}")
                                continue
                            
                            # Generate file hash to detect duplicates
                            file_hash = hashlib.md5(payload).hexdigest()
                            if file_hash in processed_files:
                                print(f"   ⚠️ Duplicate PDF: {filename}")
                                continue
                                
                            processed_files[file_hash] = True
                            
                            # Save PDF with order ID as name
                            pdf_name = f"{order_id}.pdf"
                            pdf_path = os.path.join(SAVE_DIR, pdf_name)
                            
                            with open(pdf_path, "wb") as f:
                                f.write(payload)
                            
                            print(f"   💾 Saved PDF: {pdf_name}")
                            
                            # Extract Subtotal from PDF
                            with pdfplumber.open(pdf_path) as pdf:
                                pdf_text = ""
                                for page in pdf.pages:
                                    text = page.extract_text()
                                    if text:
                                        pdf_text += text + "\n"
                            
                            amount = extract_subtotal(pdf_text)
                            
                            if amount is None:
                                print("   ⚠️ Subtotal not found in PDF")
                            else:
                                print(f"   ✓ Found Subtotal: ₹{amount:,.2f}")
                                pdf_processed = True
                                processed_orders.add(order_id)
                            
                            # Only process one PDF per email
                            break
                            
                        except Exception as e:
                            print(f"   ⚠️ Failed to process PDF: {str(e)}")
            
            # Create PDF if no valid attachment found
            if not pdf_processed:
                # Extract Subtotal from email body
                amount = extract_subtotal(cleaned_body)
                
                if amount is None:
                    # Try again with unconverted body
                    amount = extract_subtotal(body_for_processing)
                    
                if amount is None:
                    print("   ⚠️ Subtotal not found in email body - skipping")
                    continue
                
                print(f"   ✓ Found Subtotal: ₹{amount:,.2f}")
                
                # Create PDF from email body
                pdf_name = f"{order_id}.pdf"
                pdf_path = os.path.join(SAVE_DIR, pdf_name)
                
                if create_pdf_from_body(order_id, subject, body_for_pdf, pdf_path):
                    print(f"   ✓ Created PDF: {pdf_name}")
                    processed_orders.add(order_id)
                    pdf_processed = True
            
            # Add to invoice data if processed
            if pdf_processed and pdf_path:
                invoice_data.append({
                    "Order ID": order_id,
                    "Subject": subject,
                    "Amount": amount,
                    "File": pdf_path
                })
            
            processed_count += 1
            print(f"   ✔ Completed in {time.time() - start_time:.2f}s")

        except Exception as e:
            print(f"   ✖ Error processing email: {str(e)}")
            continue

    mail.logout()
    print("\n📊 Processing Summary:")
    print(f"Total emails processed: {processed_count}/{total_emails}")
    print(f"Valid invoices created: {len(invoice_data)}")
    return invoice_data

def generate_excel_report(data):
    """Generate Excel report with total sum."""
    if not data:
        print("⚠️ No valid invoices to report")
        return
        
    print("\n📊 Generating Excel report...")
    wb = Workbook()
    ws = wb.active
    ws.title = "1mg Invoices"
    
    # Headers
    headers = ["Order ID", "Subject", "Amount (₹)", "File Name"]
    ws.append(headers)
    
    # Make header row bold
    header_font = Font(bold=True)
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = header_font
    
    # Data
    total_amount = 0
    for item in data:
        amount = item.get("Amount", 0)
        if isinstance(amount, (int, float)):
            total_amount += amount
            amount_str = f"₹{amount:,.2f}"
        else:
            amount_str = "Invalid"
            
        ws.append([
            item.get("Order ID", "N/A"),
            item.get("Subject", "No Subject")[:255],
            amount_str,
            os.path.basename(item.get("File", ""))
        ])
    
    # Add total row
    ws.append([])
    ws.append(["TOTAL", "", f"₹{total_amount:,.2f}", ""])
    
    # Make total row bold
    for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(bold=True)
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                value = str(cell.value) if cell.value else ""
                if len(value) > max_length:
                    max_length = len(value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save
    report_path = os.path.join(SAVE_DIR, "1mg_Invoices_Report.xlsx")
    wb.save(report_path)
    print(f"✓ Excel report generated: {report_path}")
    print(f"💰 Total Subtotal: ₹{total_amount:,.2f}")
    return report_path

if __name__ == "__main__":
    start_time = datetime.now()
    print(f"⏳ Starting 1mg invoice processing at {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Initialize tracking sets
    processed_orders = set()
    processed_files = {}
    
    # Process emails
    invoices = fetch_1mg_emails()
    
    # Generate report
    generate_excel_report(invoices)
    
    end_time = datetime.now()
    print(f"⏱️ Total processing time: {end_time - start_time}")
    print("✅ Script completed")
